# Utilitaires pour génération de PDF et Excel

from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# GÉNÉRATION PDF - EMPLOI DU TEMPS
class EmploiDuTempsPDF:
    """
    Générateur de PDF pour emploi du temps.
    """
    
    def __init__(self, filiere, semestre, annee_academique, cours_par_jour):
        self.filiere = filiere
        self.semestre = semestre
        self.annee_academique = annee_academique
        self.cours_par_jour = cours_par_jour
        self.buffer = BytesIO()
        self.pagesize = landscape(A4)
        self.width, self.height = self.pagesize
        
    def generate(self):
        """Générer le PDF."""
        doc = SimpleDocTemplate(
            self.buffer,
            pagesize=self.pagesize,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=2*cm,
            bottomMargin=1*cm
        )
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#283593'),
            spaceAfter=12,
            alignment=TA_LEFT
        )
        
        # Contenu
        content = []
        
        # Titre
        title = f"EMPLOI DU TEMPS - {self.filiere.nom}"
        content.append(Paragraph(title, title_style))
        
        # Sous-titre
        subtitle = f"Semestre {self.semestre} - {self.annee_academique.code}"
        content.append(Paragraph(subtitle, heading_style))
        content.append(Spacer(1, 0.5*cm))
        
        # Tableau pour chaque jour
        jours = ['LUNDI', 'MARDI', 'MERCREDI', 'JEUDI', 'VENDREDI', 'SAMEDI']
        
        for jour in jours:
            cours_jour = self.cours_par_jour.get(jour, [])
            
            if not cours_jour:
                continue
            
            # En-tête du jour
            content.append(Paragraph(f"<b>{jour}</b>", heading_style))
            
            # Données du tableau
            data = [['Horaire', 'Matière', 'Type', 'Enseignant', 'Salle']]
            
            for cours in cours_jour:
                creneau = cours.creneau
                horaire = f"{creneau.heure_debut.strftime('%H:%M')} - {creneau.heure_fin.strftime('%H:%M')}"
                matiere = f"{cours.matiere.code}\n{cours.matiere.nom}"
                type_cours = cours.get_type_cours_display()
                enseignant = cours.enseignant.user.get_full_name() if cours.enseignant else "Non assigné"
                salle = cours.salle.code if cours.salle else "Non assignée"
                
                data.append([horaire, matiere, type_cours, enseignant, salle])
            
            # Créer le tableau
            table = Table(data, colWidths=[3*cm, 8*cm, 2*cm, 5*cm, 3*cm])
            
            # Style du tableau
            table.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3f51b5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Contenu
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                
                # Grille
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                
                # Alternance de couleurs
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            content.append(table)
            content.append(Spacer(1, 0.8*cm))
        
        # Pied de page
        footer = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        content.append(Spacer(1, 1*cm))
        content.append(Paragraph(footer, footer_style))
        
        # Construire le PDF
        doc.build(content)
        
        # Retourner le buffer
        self.buffer.seek(0)
        return self.buffer

# GÉNÉRATION EXCEL - EMPLOI DU TEMPS
class EmploiDuTempsExcel:
    """
    Générateur Excel pour emploi du temps.
    """
    
    def __init__(self, filiere, semestre, annee_academique, cours_par_jour):
        self.filiere = filiere
        self.semestre = semestre
        self.annee_academique = annee_academique
        self.cours_par_jour = cours_par_jour
        self.buffer = BytesIO()
        
    def generate(self):
        """Générer le fichier Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Emploi du Temps"
        
        # Styles
        title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='3F51B5', end_color='3F51B5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Titre principal
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"EMPLOI DU TEMPS - {self.filiere.nom}"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        ws.row_dimensions[1].height = 25
        
        # Sous-titre
        ws.merge_cells('A2:E2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f"Semestre {self.semestre} - {self.annee_academique.code}"
        subtitle_cell.font = Font(name='Arial', size=10, italic=True)
        subtitle_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20
        
        # Ligne vide
        current_row = 4
        
        # Pour chaque jour
        jours = ['LUNDI', 'MARDI', 'MERCREDI', 'JEUDI', 'VENDREDI', 'SAMEDI']
        
        for jour in jours:
            cours_jour = self.cours_par_jour.get(jour, [])
            
            if not cours_jour:
                continue
            
            # En-tête du jour
            ws.merge_cells(f'A{current_row}:E{current_row}')
            jour_cell = ws[f'A{current_row}']
            jour_cell.value = jour
            jour_cell.font = Font(name='Arial', size=12, bold=True)
            jour_cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            jour_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
            # En-têtes du tableau
            headers = ['Horaire', 'Matière', 'Type', 'Enseignant', 'Salle']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
            # Données
            for cours in cours_jour:
                creneau = cours.creneau
                horaire = f"{creneau.heure_debut.strftime('%H:%M')} - {creneau.heure_fin.strftime('%H:%M')}"
                matiere = f"{cours.matiere.code}\n{cours.matiere.nom}"
                type_cours = cours.get_type_cours_display()
                enseignant = cours.enseignant.user.get_full_name() if cours.enseignant else "Non assigné"
                salle = cours.salle.code if cours.salle else "Non assignée"
                
                row_data = [horaire, matiere, type_cours, enseignant, salle]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.alignment = cell_alignment
                    cell.border = border
                
                ws.row_dimensions[current_row].height = 30
                current_row += 1
            
            # Ligne vide entre les jours
            current_row += 1
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        
        # Pied de page
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        footer_cell = ws[f'A{current_row}']
        footer_cell.value = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        footer_cell.font = Font(name='Arial', size=8, italic=True, color='808080')
        footer_cell.alignment = Alignment(horizontal='center')
        
        # Sauvegarder
        wb.save(self.buffer)
        self.buffer.seek(0)
        return self.buffer

# GÉNÉRATION PDF - LISTE DES CONFLITS
class ConflitsPDF:
    """
    Générateur de PDF pour liste des conflits.
    """
    
    def __init__(self, conflits, annee_academique):
        self.conflits = conflits
        self.annee_academique = annee_academique
        self.buffer = BytesIO()
        
    def generate(self):
        """Générer le PDF."""
        doc = SimpleDocTemplate(
            self.buffer,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=2*cm,
            bottomMargin=1*cm
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#c62828'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Contenu
        content = []
        
        # Titre
        title = f"RAPPORT DES CONFLITS D'EMPLOI DU TEMPS"
        content.append(Paragraph(title, title_style))
        
        subtitle = f"Année académique : {self.annee_academique.code}"
        content.append(Paragraph(subtitle, styles['Heading2']))
        content.append(Spacer(1, 0.5*cm))
        
        # Statistiques
        total = len(self.conflits)
        non_resolus = sum(1 for c in self.conflits if c.statut in ['DETECTE', 'EN_COURS'])
        
        stats = f"<b>Total de conflits :</b> {total} | <b>Non résolus :</b> {non_resolus}"
        content.append(Paragraph(stats, styles['Normal']))
        content.append(Spacer(1, 0.5*cm))
        
        # Tableau des conflits
        data = [['Type', 'Statut', 'Cours 1', 'Cours 2', 'Description', 'Date']]
        
        for conflit in self.conflits:
            type_conflit = conflit.get_type_conflit_display()
            statut = conflit.get_statut_display()
            
            cours1_info = f"{conflit.cours1.matiere.code}\n{conflit.cours1.filiere.code}"
            
            if conflit.cours2:
                cours2_info = f"{conflit.cours2.matiere.code}\n{conflit.cours2.filiere.code}"
            else:
                cours2_info = "N/A"
            
            description = conflit.description[:50] + "..." if len(conflit.description) > 50 else conflit.description
            date = conflit.date_detection.strftime('%d/%m/%Y')
            
            data.append([type_conflit, statut, cours1_info, cours2_info, description, date])
        
        # Créer le tableau
        table = Table(data, colWidths=[3*cm, 2.5*cm, 4*cm, 4*cm, 7*cm, 2.5*cm])
        
        # Style
        table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c62828')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Contenu
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            
            # Grille
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            
            # Alternance
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        content.append(table)
        
        # Footer
        content.append(Spacer(1, 1*cm))
        footer = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        content.append(Paragraph(footer, styles['Normal']))
        
        # Build
        doc.build(content)
        
        self.buffer.seek(0)
        return self.buffer

# GÉNÉRATION PDF - PLANNING ENSEIGNANT
class PlanningEnseignantPDF:
    """
    Générateur de PDF pour planning enseignant.
    """
    
    def __init__(self, enseignant, cours_par_jour, annee_academique):
        self.enseignant = enseignant
        self.cours_par_jour = cours_par_jour
        self.annee_academique = annee_academique
        self.buffer = BytesIO()
        
    def generate(self):
        """Générer le PDF."""
        doc = SimpleDocTemplate(
            self.buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        # Contenu
        content = []
        
        # Titre
        title = f"PLANNING DE {self.enseignant.user.get_full_name()}"
        content.append(Paragraph(title, title_style))
        
        subtitle = f"{self.enseignant.get_grade_display()} - {self.annee_academique.code}"
        content.append(Paragraph(subtitle, styles['Heading3']))
        content.append(Spacer(1, 0.5*cm))
        
        # Tableau par jour
        jours = ['LUNDI', 'MARDI', 'MERCREDI', 'JEUDI', 'VENDREDI', 'SAMEDI']
        
        for jour in jours:
            cours_jour = self.cours_par_jour.get(jour, [])
            
            if not cours_jour:
                continue
            
            content.append(Paragraph(f"<b>{jour}</b>", styles['Heading2']))
            
            data = [['Horaire', 'Matière', 'Filière', 'Salle']]
            
            for cours in cours_jour:
                horaire = f"{cours.creneau.heure_debut.strftime('%H:%M')} - {cours.creneau.heure_fin.strftime('%H:%M')}"
                matiere = f"{cours.matiere.code} - {cours.matiere.nom}"
                filiere = cours.filiere.code
                salle = cours.salle.code if cours.salle else "N/A"
                
                data.append([horaire, matiere, filiere, salle])
            
            table = Table(data, colWidths=[3*cm, 8*cm, 3*cm, 3*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3f51b5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            content.append(table)
            content.append(Spacer(1, 0.5*cm))
        
        # Footer
        footer = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        content.append(Paragraph(footer, styles['Normal']))
        
        doc.build(content)
        
        self.buffer.seek(0)
        return self.buffer
